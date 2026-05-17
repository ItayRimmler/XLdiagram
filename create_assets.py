"""
Generates template.xlsx and example.xlsx in inputs/.
Run once: py create_assets.py
"""

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

INPUTS = Path(__file__).parent / "inputs"
INPUTS.mkdir(exist_ok=True)

# ---------------------------------------------------------------------------
# 50 draw.io styles  (group name will be prefixed with # in template)
# ---------------------------------------------------------------------------
STYLES = [
    # ── Basic shapes ────────────────────────────────────────────────────────
    ("Rectangle",           120, 60,  "rounded=0;whiteSpace=wrap;html=1;"),
    ("Rounded",             120, 60,  "rounded=1;whiteSpace=wrap;html=1;"),
    ("Ellipse",             120, 60,  "ellipse;whiteSpace=wrap;html=1;"),
    ("Diamond",             120, 80,  "rhombus;whiteSpace=wrap;html=1;"),
    ("Parallelogram",       120, 60,  "shape=parallelogram;perimeter=parallelogramPerimeter;whiteSpace=wrap;html=1;"),
    ("Hexagon",             120, 60,  "shape=hexagon;perimeter=hexagonPerimeter2;whiteSpace=wrap;html=1;"),
    ("Triangle",            100, 80,  "triangle;whiteSpace=wrap;html=1;"),
    ("Cylinder",            80,  100, "shape=mxgraph.flowchart.stored_data;whiteSpace=wrap;html=1;"),
    ("Note",                120, 80,  "shape=note;whiteSpace=wrap;html=1;backgroundOutline=1;"),
    ("Terminator",          120, 50,  "rounded=1;whiteSpace=wrap;html=1;arcSize=50;"),
    # ── Coloured rectangles ─────────────────────────────────────────────────
    ("Blue Box",            120, 60,  "rounded=0;whiteSpace=wrap;html=1;fillColor=#dae8fc;strokeColor=#6c8ebf;"),
    ("Green Box",           120, 60,  "rounded=0;whiteSpace=wrap;html=1;fillColor=#d5e8d4;strokeColor=#82b366;"),
    ("Red Box",             120, 60,  "rounded=0;whiteSpace=wrap;html=1;fillColor=#f8cecc;strokeColor=#b85450;"),
    ("Yellow Box",          120, 60,  "rounded=0;whiteSpace=wrap;html=1;fillColor=#fff2cc;strokeColor=#d6b656;"),
    ("Orange Box",          120, 60,  "rounded=0;whiteSpace=wrap;html=1;fillColor=#ffe6cc;strokeColor=#d79b00;"),
    ("Purple Box",          120, 60,  "rounded=0;whiteSpace=wrap;html=1;fillColor=#e1d5e7;strokeColor=#9673a6;"),
    ("Gray Box",            120, 60,  "rounded=0;whiteSpace=wrap;html=1;fillColor=#f5f5f5;strokeColor=#666666;fontColor=#333333;"),
    ("Dark Box",            120, 60,  "rounded=0;whiteSpace=wrap;html=1;fillColor=#1e4d78;strokeColor=#1e4d78;fontColor=#ffffff;"),
    ("Teal Box",            120, 60,  "rounded=0;whiteSpace=wrap;html=1;fillColor=#d5ffd5;strokeColor=#009999;"),
    ("Pink Box",            120, 60,  "rounded=0;whiteSpace=wrap;html=1;fillColor=#ffd5e0;strokeColor=#cc6699;"),
    # ── Coloured rounded ────────────────────────────────────────────────────
    ("Blue Rounded",        120, 60,  "rounded=1;whiteSpace=wrap;html=1;fillColor=#dae8fc;strokeColor=#6c8ebf;"),
    ("Green Rounded",       120, 60,  "rounded=1;whiteSpace=wrap;html=1;fillColor=#d5e8d4;strokeColor=#82b366;"),
    ("Red Rounded",         120, 60,  "rounded=1;whiteSpace=wrap;html=1;fillColor=#f8cecc;strokeColor=#b85450;"),
    ("Yellow Rounded",      120, 60,  "rounded=1;whiteSpace=wrap;html=1;fillColor=#fff2cc;strokeColor=#d6b656;"),
    ("Orange Rounded",      120, 60,  "rounded=1;whiteSpace=wrap;html=1;fillColor=#ffe6cc;strokeColor=#d79b00;"),
    ("Purple Rounded",      120, 60,  "rounded=1;whiteSpace=wrap;html=1;fillColor=#e1d5e7;strokeColor=#9673a6;"),
    # ── Flowchart ───────────────────────────────────────────────────────────
    ("FC Process",          120, 60,  "shape=mxgraph.flowchart.process;whiteSpace=wrap;html=1;"),
    ("FC Decision",         100, 80,  "shape=mxgraph.flowchart.decision;whiteSpace=wrap;html=1;"),
    ("FC Database",         80,  100, "shape=mxgraph.flowchart.database;whiteSpace=wrap;html=1;"),
    ("FC Document",         120, 70,  "shape=mxgraph.flowchart.document;whiteSpace=wrap;html=1;"),
    ("FC Manual Input",     120, 60,  "shape=mxgraph.flowchart.manual_input;whiteSpace=wrap;html=1;"),
    ("FC Delay",            120, 60,  "shape=mxgraph.flowchart.delay;whiteSpace=wrap;html=1;"),
    ("FC Display",          120, 60,  "shape=mxgraph.flowchart.display;whiteSpace=wrap;html=1;"),
    ("FC Preparation",      120, 60,  "shape=mxgraph.flowchart.preparation;whiteSpace=wrap;html=1;"),
    ("FC Predefined",       120, 60,  "shape=mxgraph.flowchart.predefined_process;whiteSpace=wrap;html=1;"),
    # ── Network / IT ────────────────────────────────────────────────────────
    ("Server",              80,  80,  "shape=mxgraph.network.server;html=1;whiteSpace=wrap;"),
    ("Database Server",     80,  80,  "shape=mxgraph.network.database;html=1;whiteSpace=wrap;"),
    ("Switch",              80,  80,  "shape=mxgraph.network.switch_2;html=1;whiteSpace=wrap;"),
    ("Router",              80,  80,  "shape=mxgraph.network.router;html=1;whiteSpace=wrap;"),
    ("Firewall",            80,  80,  "shape=mxgraph.network.firewall;html=1;whiteSpace=wrap;"),
    ("Laptop",              80,  80,  "shape=mxgraph.network.laptop;html=1;whiteSpace=wrap;"),
    ("Desktop",             80,  80,  "shape=mxgraph.network.workstation;html=1;whiteSpace=wrap;"),
    ("Cloud",               100, 80,  "shape=mxgraph.network.cloud;html=1;whiteSpace=wrap;"),
    ("Mobile",              60,  100, "shape=mxgraph.android.phone2;html=1;whiteSpace=wrap;"),
    # ── Styles / borders ────────────────────────────────────────────────────
    ("Dashed Border",       120, 60,  "rounded=0;whiteSpace=wrap;html=1;dashed=1;"),
    ("Thick Border",        120, 60,  "rounded=0;whiteSpace=wrap;html=1;strokeWidth=3;"),
    ("Outline Only",        120, 60,  "rounded=0;whiteSpace=wrap;html=1;fillColor=none;"),
    ("Double Border",       120, 60,  "rounded=0;whiteSpace=wrap;html=1;double=1;"),
    ("Shadow",              120, 60,  "rounded=0;whiteSpace=wrap;html=1;shadow=1;"),
    ("Bold Dark",           120, 60,  "rounded=0;whiteSpace=wrap;html=1;strokeWidth=2;fillColor=#333333;fontColor=#ffffff;strokeColor=#333333;"),
]

assert len(STYLES) == 50, f"Expected 50 styles, got {len(STYLES)}"


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def _header_font():
    return Font(bold=True)

def _header_fill():
    return PatternFill("solid", fgColor="BDD7EE")

def _comment_fill():
    return PatternFill("solid", fgColor="F2F2F2")

def _thin_border():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

def _write_headers(ws, headers):
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = _header_font()
        c.fill = _header_fill()
        c.alignment = Alignment(horizontal="center")


# ---------------------------------------------------------------------------
# template.xlsx
# ---------------------------------------------------------------------------
def make_template():
    wb = Workbook()

    # ── Layout sheet ────────────────────────────────────────────────────────
    ws_layout = wb.active
    ws_layout.title = "Layout"
    _write_headers(ws_layout, ["Node name (or Name|Group)"])
    ws_layout.column_dimensions["A"].width = 28

    # ── Edges sheet ─────────────────────────────────────────────────────────
    ws_edges = wb.create_sheet("Edges")
    _write_headers(ws_edges, ["Source", "Separator", "Target", "Label (optional)"])
    for col, w in zip("ABCD", [20, 12, 20, 20]):
        ws_edges.column_dimensions[col].width = w
    # separator reference rows (grayed out)
    sep_ref = [
        ("", ">",  "", "forward arrow"),
        ("", "<",  "", "backward arrow"),
        ("", "o>", "", "dot → arrow"),
        ("", "<o", "", "arrow → dot"),
        ("", "<>", "", "bidirectional"),
        ("", "-",  "", "plain wire"),
        ("", "o-", "", "dot → line"),
        ("", "-o", "", "line → dot"),
    ]
    for i, row in enumerate(sep_ref, 2):
        for col, val in enumerate(row, 1):
            c = ws_edges.cell(row=i, column=col, value=val)
            c.font = Font(italic=True, color="808080")

    # ── Style sheet ─────────────────────────────────────────────────────────
    ws_style = wb.create_sheet("Style")
    _write_headers(ws_style, ["# = off", "Group", "Width", "Height", "draw.io Style String"])
    for col, w in zip("ABCDE", [8, 24, 10, 10, 80]):
        ws_style.column_dimensions[col].width = w

    for i, (name, w, h, style) in enumerate(STYLES, 2):
        ws_style.cell(row=i, column=1, value="#").fill = _comment_fill()
        ws_style.cell(row=i, column=2, value=name).fill = _comment_fill()
        ws_style.cell(row=i, column=3, value=w).fill = _comment_fill()
        ws_style.cell(row=i, column=4, value=h).fill = _comment_fill()
        ws_style.cell(row=i, column=5, value=style).fill = _comment_fill()

    wb.save(INPUTS / "template.xlsx")
    print("Written: inputs/template.xlsx")


# ---------------------------------------------------------------------------
# example.xlsx  — small network diagram
# Groups match STYLES names exactly so users can see how to activate a style.
# ---------------------------------------------------------------------------
EXAMPLE_LAYOUT = {
    # (row, col): "NodeName|Group"   — Group must match a Style sheet group name
    (2, 3): "Internet|Cloud",
    (4, 2): "FW|Firewall",
    (4, 4): "RTR|Router",
    (6, 1): "PC-1|Desktop",
    (6, 2): "PC-2|Desktop",
    (6, 3): "SW|Switch",
    (6, 4): "SRV-1|Server",
    (6, 5): "SRV-2|Server",
}

EXAMPLE_EDGES = [
    ("Internet", ">",  "FW",    "WAN"),
    ("Internet", ">",  "RTR",   "WAN"),
    ("FW",       ">",  "SW",    ""),
    ("RTR",      ">",  "SW",    ""),
    ("SW",       ">",  "PC-1",  ""),
    ("SW",       ">",  "PC-2",  ""),
    ("SW",       ">",  "SRV-1", ""),
    ("SW",       ">",  "SRV-2", ""),
]

# Styles used in the example — drawn from STYLES list above.
# All other 50 styles are copied from STYLES with # still on.
EXAMPLE_ACTIVE = {"Cloud", "Firewall", "Router", "Switch", "Server", "Desktop"}


def make_example():
    wb = Workbook()

    # Layout
    ws_layout = wb.active
    ws_layout.title = "Layout"
    _write_headers(ws_layout, ["Node name (or Name|Group)"])
    ws_layout.column_dimensions["A"].width = 24
    for (row, col), name in EXAMPLE_LAYOUT.items():
        ws_layout.cell(row=row, column=col, value=name)

    # Edges
    ws_edges = wb.create_sheet("Edges")
    _write_headers(ws_edges, ["Source", "Separator", "Target", "Label (optional)"])
    for col, w in zip("ABCD", [20, 12, 20, 20]):
        ws_edges.column_dimensions[col].width = w
    for i, (src, sep, tgt, lbl) in enumerate(EXAMPLE_EDGES, 2):
        ws_edges.cell(row=i, column=1, value=src)
        ws_edges.cell(row=i, column=2, value=sep)
        ws_edges.cell(row=i, column=3, value=tgt)
        ws_edges.cell(row=i, column=4, value=lbl)

    # Style — all 50 from STYLES; active ones have blank col A, inactive keep "#"
    ws_style = wb.create_sheet("Style")
    _write_headers(ws_style, ["# = off", "Group", "Width", "Height", "draw.io Style String"])
    for col, w in zip("ABCDE", [8, 24, 10, 10, 80]):
        ws_style.column_dimensions[col].width = w
    for i, (name, w, h, style) in enumerate(STYLES, 2):
        active_flag = "" if name in EXAMPLE_ACTIVE else "#"
        fill = _comment_fill() if active_flag == "#" else None
        for col, val in enumerate([active_flag, name, w, h, style], 1):
            c = ws_style.cell(row=i, column=col, value=val)
            if fill:
                c.fill = fill

    wb.save(INPUTS / "example.xlsx")
    print("Written: inputs/example.xlsx")


if __name__ == "__main__":
    make_template()
    make_example()
